from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("demo_colors.xlsx")
ws = wb.active

for cell in ws['C']:
    print(cell.value)
    if cell.value != "CSAT":
        if float(cell.value) <= 4.0:
            cell.fill = PatternFill("solid", start_color="FFA500")
        if float(cell.value) >= 4.0:
            cell.fill = PatternFill("solid", start_color="00FF00")

wb.save('demo_colors_updated.xlsx')

